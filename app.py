import math
import os
import time
import zipfile
import openpyxl

from flask import Flask, jsonify, send_from_directory, send_file
from flask_cors import CORS
from flask_socketio import SocketIO
from dataclasses import asdict
from io import BytesIO

from models.ShopItem import ShopItem
from routers import routers
from utills.SingletonWebDriver import SingletonWebDriver
from utills.open_chrome import open_chrome, close_chrome
from utills.save_image_and_return_abs_path import save_image_and_return_abs_path
from utills.print_running_time import print_running_time
from utills.convert_sets_to_lists import convert_sets_to_lists
from services import market_search
from services import taobao_search
from services.pruning_shop_item import selling_keywords_item, search_sub_keywords, pruning_naver_shoping

app = Flask(__name__, static_folder='../utopia_frontend/dist')
CORS(app, origins=["http://127.0.0.1:5000"])
socketio = SocketIO(app, cors_allowed_origins="http://127.0.0.1:5000")
EXCEL_PERCENTY_PATH = 'percenty.xlsx'


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, 'index.html')


@app.route('/test-signal')
def test_sig_func():
    return routers.test_signal()


@app.route('/debug-test')
def test_debug():


@app.route('/api/check-item')
def check_item():
    chrome_process = open_chrome()

    driver = SingletonWebDriver.get_driver()
    search_platform = 'auction'
    keywordlist, min_price, max_price, max_cnt_item, is_pruning = routers.process_request()
    is_auto_flag = False
    print(keywordlist)

    selling_items = source_items(driver, keywordlist, min_price, max_price, max_cnt_item, is_auto_flag, search_platform)

    close_chrome(chrome_process)
    print('done.')
    return jsonify(selling_items)


@app.route('/api/pruning-items')
def pruning_item():
    chrome_process = open_chrome()
    driver = SingletonWebDriver.get_driver()
    item_name, is_pruning = routers.pruning_request()

    selling_keywords = [
        ' '.join(name.split()[:3]) for name in item_name
    ]

    # 타오바오 로그인
    try:
        taobao_search.taobao_login(driver)
    except Exception as e:
        print("err:", e)
        SingletonWebDriver.close_driver()
        return e

    shop_items = prune_items(driver, selling_keywords, is_pruning)

    shop_items_list, cnt = taobao_searching_items(driver, 300, shop_items)
    print(shop_items_list)

    SingletonWebDriver.close_driver()
    close_chrome(chrome_process)

    print('done.')
    return shop_items_list


@app.route('/api/market-search')
def market_search_func():
    chrome_process = open_chrome()

    driver = SingletonWebDriver.get_driver()
    search_platform = 'auction'
    keywordlist, min_price, max_price, max_cnt_item, is_pruning = routers.process_request()
    is_auto_flag = True
    print(min_price, max_price, max_cnt_item, is_pruning)
    print(keywordlist)

    # 수집 수 보정 값   //보정치 = 5
    calibration_max_cnt = math.ceil(max_cnt_item)
    print('보정치 : ', calibration_max_cnt)

    # 소켓 시작 시간 전송
    socketio.emit('message', {
        'status': 'start',
        'processed': 1,
        'total': 5,
        'infoMsg': '타오바오 로그인이 필요합니다.'
    })

    # 타오바오 로그인
    try:
        taobao_search.taobao_login(driver)
    except Exception as e:
        print("err:", e)
        SingletonWebDriver.close_driver()
        return e
    start_time = time.localtime()
    print(f'수집 시작 시간: {start_time.tm_hour}:{start_time.tm_min}:{start_time.tm_sec}')

    # 아이템 소싱 파트
    selling_keywords = source_items(driver, keywordlist, min_price, max_price, calibration_max_cnt, is_auto_flag, search_platform)

    # 아이템 가지치기 파트
    shop_items = prune_items(driver, selling_keywords, is_pruning)

    # 타오바오 링크, 타오바오 이미지 수집
    res_data, cnt = taobao_searching_items(driver, 300, shop_items)
    print(res_data)

    end_time = time.localtime()
    print(f'수집 종료 시간: {end_time.tm_hour}:{end_time.tm_min}:{end_time.tm_sec}')
    SingletonWebDriver.close_driver()
    close_chrome(chrome_process)

    work_time = print_running_time(start_time, end_time)
    socketio.emit('message', {
        'status': 'end',
        'processed': 5,
        'total': 5,
        'time': work_time,
        'infoMsg': '완료'
    })
    print(f'수집 아이템 수: {cnt}개')

    shop_items_dict = [asdict(item) for item in res_data]
    shop_items_list = convert_sets_to_lists(shop_items_dict)
    print('done.')
    return jsonify(shop_items_list)


def source_items(driver, keywordlist, min_price, max_price, calibration_max_cnt, is_auto_flag, search_platform):
    selling_keywords = []

    while True:
        # keywordlist가 3개 이상이면 앞에서 2개만 사용해서 검색
        if len(keywordlist) >= 2:
            current_keywords = keywordlist[:2]
        else:
            current_keywords = keywordlist  # keywordlist가 2개 미만인 경우 전체 사용

        # 마켓 검색
        shop_list = market_search.search_shops(driver, current_keywords, search_platform)
        print(shop_list)
        socketio.emit('message', {
            'status': 'in_progress',
            'processed': 2,
            'total': 5,
            'infoMsg': f'{len(shop_list)}개의 마켓을 찾음'
        })
        print(f'{len(shop_list)}개의 마켓을 찾음')

        # 쿠키 저장 및 재사용
        cookies = driver.get_cookies()
        for cookie in cookies:
            driver.add_cookie(cookie)

        # 팔린 키워드 추출
        new_selling_keywords = selling_keywords_item(driver, shop_list, min_price, max_price, calibration_max_cnt, is_auto_flag, search_platform)
        selling_keywords.extend(new_selling_keywords)
        socketio.emit('message', {
            'status': 'in_progress',
            'processed': 3,
            'total': 5,
            'infoMsg': f'{len(new_selling_keywords)}개의 아이템을 추가했습니다.'
        })

        print(f'{len(new_selling_keywords)}개의 새로운 아이템 추가, 전체 아이템: {len(selling_keywords)}개')

        # 아이템 수가 calibration_max_cnt 도달했을 경우 종료
        if len(selling_keywords) >= calibration_max_cnt:
            print("완료 - 목표 아이템 수 도달")
            break

        # 키워드 리스트에서 앞의 두 개를 제거하여 다음 키워드로 이동
        if len(keywordlist) > 2:
            keywordlist = keywordlist[2:]
        else:
            print("더 이상 사용할 키워드가 없습니다.")
            break  # 키워드가 더 이상 없으면 종료

    return selling_keywords


def prune_items(driver, selling_keywords, is_pruning):
    shop_items = []
    # 가지치기
    for item_keyword in selling_keywords:
        item = pruning_naver_shoping(driver, item_keyword, is_pruning)
        if item:
            shop_items.extend(item)

    # 서브 키워드 수집
    for item in shop_items:
        main_keyword = item.item_main_keywords
        sub_keywords, recommended_keywords = search_sub_keywords(driver, main_keyword)
        item.item_sub_keywords = sub_keywords
        item.item_recommended_keywords = recommended_keywords

    return shop_items


def taobao_searching_items(driver, max_cnt_item, shop_items):
    taobao_url = 'https://s.taobao.com/search?q='
    cnt = 0
    if max_cnt_item > len(shop_items):
        all_item_cnt = len(shop_items)
    else:
        all_item_cnt = max_cnt_item

    driver.get(taobao_url)
    socketio.emit('message', {
        'status': 'in_progress',
        'processed': 4,
        'total': 5,
        'infoMsg': '타오바오 이미지 검색 중'
    })
    res_data = []
    for item in shop_items:
        image_path = save_image_and_return_abs_path(item.item_image_url)
        if image_path:
            taobao_item = taobao_search.taobao_image_search(driver, image_path)
            if taobao_item:
                taobao_item_link, taobao_image_link = taobao_item
            else:
                continue

            print(f'{cnt + 1} / {all_item_cnt}')
            item.item_taobao_item_url = taobao_item_link
            item.item_taobao_image_url = taobao_image_link
        res_data.append(item)
        cnt += 1
        if cnt >= max_cnt_item:
            break
        time.sleep(2)

    print(f'누락 {all_item_cnt - cnt}개')
    return res_data, cnt


@app.route('/api/xlsx-convert', methods=['POST'])
def xlsx_convert():
    convert_type_code, data = routers.xlsx_data_request()
    print(convert_type_code)

    if convert_type_code == 1:
        if not os.path.exists(EXCEL_PERCENTY_PATH):
            return "Error: percenty.xlsx 파일이 없습니다.", 400

        # 데이터를 50개씩 분할
        chunk_size = 50
        data_chunks = [data[i:i + chunk_size] for i in range(0, len(data), chunk_size)]

        if len(data_chunks) == 1:
            # 데이터가 한 개 chunk인 경우, ZIP이 아닌 단일 엑셀 파일로 반환
            workbook = openpyxl.load_workbook(EXCEL_PERCENTY_PATH)

            if 'multi_ss' not in workbook.sheetnames:
                return "Error: 'multi_ss' 시트가 없습니다.", 400

            sheet = workbook['multi_ss']

            # A4 셀부터 데이터 작성
            start_row = 4
            start_col = 1

            for row_index, row_data in enumerate(data_chunks[0]):
                for col_index, value in enumerate(row_data):
                    if col_index == 4:
                        value = int(value) if value else 0  # 정수 변환 및 null 값 처리

                    cell = sheet.cell(row=start_row + row_index, column=start_col + col_index, value=value)
                    if isinstance(value, int):
                        cell.number_format = '0'

            # 메모리에서 엑셀 파일로 저장
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # 단일 엑셀 파일로 반환
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name='completed_percenty.xlsx')

        else:
            # 여러 chunk인 경우 ZIP 파일로 압축
            zip_buffer = BytesIO()

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for chunk_index, chunk_data in enumerate(data_chunks):
                    workbook = openpyxl.load_workbook(EXCEL_PERCENTY_PATH)

                    if 'multi_ss' not in workbook.sheetnames:
                        return "Error: 'multi_ss' 시트가 없습니다.", 400

                    sheet = workbook['multi_ss']

                    # A4 셀부터 데이터 작성
                    start_row = 4
                    start_col = 1

                    for row_index, row_data in enumerate(chunk_data):
                        for col_index, value in enumerate(row_data):
                            if col_index == 4:
                                value = int(value) if value else 0

                            cell = sheet.cell(row=start_row + row_index, column=start_col + col_index, value=value)
                            if isinstance(value, int):
                                cell.number_format = '0'

                    # 메모리에서 엑셀 파일로 저장
                    output = BytesIO()
                    workbook.save(output)
                    output.seek(0)

                    # 엑셀 파일을 zip 파일에 추가
                    file_name = f'completed_percenty_part_{chunk_index + 1}.xlsx'
                    zip_file.writestr(file_name, output.getvalue())

            # zip 파일 완료
            zip_buffer.seek(0)

            # 압축된 zip 파일을 전송
            return send_file(zip_buffer, mimetype='application/zip', as_attachment=True,
                             download_name='percenty_files.zip')

    elif convert_type_code == 2:
        print('미구현')
        return "미구현", 400
    elif convert_type_code == 3:
        print('미구현')
        return "미구현", 400
    else:
        return "올바른 변환법이 지정되지 않았습니다.", 400


if __name__ == '__main__':
    app.run(debug=True)
